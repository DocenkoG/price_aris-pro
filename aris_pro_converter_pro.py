# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import io
import sys
import configparser
import time
import openpyxl                   # Для .xlsx
#import xlrd                         # для .xls
from  price_tools import config_read, getCell, quoted, dump_cell, currencyTypeXlsx, subInParentheses, getCellXlsx
import price_tools
import csv



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена') :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyTypeXlsx(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert2csv_pro( pFileName):
    make_loger()
    log.debug('Begin ' + __name__ + ' convert2csv')

    # Прочитать конфигурацию из файла
    cfg = config_read( 'aris_pro_cfg_pro.cfg' )
    csvFName  = cfg.get('basic','filename_out')
    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)

    log.debug('Открываю файл '+ pFileName)
    book = openpyxl.load_workbook(filename = pFileName, read_only=False, keep_vba=False, data_only=False)
#   book = xlrd.open_workbook( pFileName.encode('cp1251'), formatting_info=True)
#   book = xlrd.open_workbook( os.path.join( mydir, pFileName.encode('cp1251')), formatting_info=True)
        
    outFile = open( csvFName, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriter = csv.DictWriter(outFile, fieldnames=out_cols )
    csvWriter.writeheader()
    ssss = []
    line_qty = 0
#   for sh in book.sheets() :
    for SheetName in book.get_sheet_names():                        # Организую цикл по страницам
        if SheetName in ('Световое оборудование') :                 # пропускаю ненужные страницы
            continue
        log.debug('Устанавливаю страницу ' + SheetName )
        #log.debug('На странице %d строк' % sh.nrows)
        log.debug('На странице %d строк' % book[SheetName].max_row)
                                                                    # цикл по строкам страницы
        sh = book[SheetName]                                        # xlsx   
        grpName = SheetName
        subGrpName= ''
        brand = ''
        recOut  ={}
        '''                                   # Блок проверки свойств для распознавания групп XLSX                                  
        for i in range(52, 58):                                                         
            i_last = i
            ccc = sh.cell( row=i, column=2 )
            print(i, ccc.value)
            print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, ccc.font.color, '------', ccc.fill.fgColor.index)
            print('------')
        '''
        #for i in range(1, sheet.nrows) :                                           # xls
        for i in range(1, sh.max_row +1) :                                       # xlsx
            i_last = i
            try:
                ccc = sh.cell( row=i, column=in_cols_j['подгруппа'] )
                impValues = getXlsxString(sh, i, in_cols_j)
                #impValues = getXlsString(sheet, i, in_cols_j)
                #print( impValues['закупка'])
                if  ccc.fill.fgColor.index == 48 :                                  # Подгруппа
                    subGrpNameNQ = sh.cell(row=i, column=in_cols_j['подгруппа']).value
                    subGrpName = quoted(subGrpNameNQ)
                    brand = subInParentheses( subGrpName)
                elif impValues['цена']=='0': # (ccc.value == None) or (ccc2.value == None) :    # Пустая строка
                    pass
                    #print( 'Пустая строка. i=',i, impValues )
                else :  
                    #impValues['валюта'] = currencyTypeXlsx(sh, i, in_cols_j['валюта']-1)
                    impValues['бренд'] = brand
                    impValues['группа_'] = grpName
                    impValues['подгруппа'] = subGrpName
                    for outColName in out_template.keys() :
                        shablon = out_template[outColName]
                        for key in impValues.keys():
                            if shablon.find(key) >= 0 :
                                shablon = shablon.replace(key, impValues[key])
                        if (outColName == 'закупка') and ('*' in shablon) :
                            vvv1 = float( shablon[ :shablon.find('*')     ] )
                            vvv2 = float( shablon[  shablon.find('*')+1:  ] )
                            #print(vvvv)
                            shablon = str( float(vvv1) * float(vvv2) )
                        recOut[outColName] = shablon

                    csvWriter.writerow(recOut)

            except Exception as e:
                print(e)
                if str(e) == "'NoneType' object has no attribute 'rgb'":
                    pass
                else:
                    log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )

    log.info('Обработано ' +str(i_last)+ ' строк.')
    outFile.close()
