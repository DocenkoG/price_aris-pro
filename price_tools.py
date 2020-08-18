# -*- coding: UTF-8 -*-
import xlrd                         # для .xls
import openpyxl                     # Для .xlsx
import re
import configparser
import os.path



def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):     
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):     
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    if price_datetime+qty_seconds < time.time() :
        file_age = round((time.time()-price_datetime)/24/60/60)
        log.error('Файл "'+fileName+'" устарел!  Допустимый период '+ str(qty_days)+' дней, а ему ' + str(file_age) )
        return False
    else:
        return True



def openX(fileName ):
    typeX = fileName[fileName.find('.')+1 :]
    if typeX.lower() == 'xlsx':
        book = openpyxl.load_workbook(filename = fileName, read_only=False, keep_vba=False, data_only=False) # xlsx
    else:
        book = xlrd.open_workbook( fileName.encode('cp1251'), formatting_info=True)                          # xls
    return book



def sheetByName( fileName
                ,sheetName):
    typeX = fileName[fileName.find('.')+1 :]
    try:
        if typeX.lower() == 'xlsx':
            book = openpyxl.load_workbook(filename = fileName, read_only=False, keep_vba=False, data_only=False) # xlsx
            sheet = book[sheetName]                                                                              # xlsx 
        else:
            book = xlrd.open_workbook( fileName.encode('cp1251'), formatting_info=True)                          # xls
            sheet = book.sheet_by_name(sheetName)
    except Exception as e:
        print("<%s> <%s> <%S> <<%s>>",fileName, typeX, sheetName,e)
        sheet = False
    return sheet

    #sheet = book.worksheets[0]                                                                              # xlsx
    #sheet = book.sheets()[0]                                                                                # xls

def getCellXlsx(  row       # номер строки
                , col       # номер колонки 
                , isDigit   # Признак, числовое ли значение нужно из этого поля
                , sheet     #  лист XLSX
                ):
    '''
    Функция возвращает значение xls-ячейки в виде строки.    
    Для цифровых ячеек делается предварительное преобразование 
    в число (пустые и нечисловые значения преобразуются в "0")
    '''
    ccc = sheet.cell(row=row, column=col)
    cellType  = ccc.data_type
    cellValue = ccc.value
    if (isDigit == 'Y') : 
        if (cellValue == None) : 
            ss = '0'
        elif (cellType in ('n')) :                  # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
#           ss = '0'
            try:
                ss = str(float(cellValue.replace(',','.')))
            except ValueError as e:
                ss='0'
    else :
        if (cellValue == None) : 
            ss = ''
        elif (cellType in ('n')) :                    # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
            ss = str(cellValue)
    return ss



def getCell(  row       # номер строки
            , col       # номер колонки 
            , isDigit   # Признак, числовое ли значение нужно из этого поля
            , sheet     #  лист XLS
            ):
    '''
    Функция возвращает значение xls-ячейки в виде строки.    
    Для цифровых ячеек делается предварительное преобразование 
    в число (пустые и нечисловые значения преобразуются в "0")
    '''
    ccc = sheet.cell(row, col)
    cellType  = ccc.ctype
    cellValue = ccc.value
    if (isDigit == 'Y') : 
        if (cellValue == '') : 
            ss = '0'
        elif (cellType in (2,3)) :                  # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
            if cellValue == 'по запросу':
                ss = '0.1'
            else:
                ss = '0'
    else :
        if (cellType in (2,3)) :                    # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
            ss = str(cellValue)
    return ss



def subInParentheses( sourceString):
    re_parentheses = re.compile('^.*\(([^)]*)\).*$',  re.IGNORECASE )
    is_parentheses = re_parentheses.match(sourceString)
    if is_parentheses:                        # Файл соответствует шаблону имени
        key = is_parentheses.group(1)         # выделяю ключ из имени файла
    else:
        key = '' 
    return key



def currencyTypeXlsx(row, col, sheet):
    '''
    Функция анализирует "формат ячейки" таблицы excel, является ли он "денежным"
    и какая валюта указана в этом формате.
    Распознаются не все валюты и способы их описания.
    '''
    c = sheet.cell(row=row, column=col)
    fmt_str = c.number_format
    if '\u20bd' in fmt_str:
        val = 'RUB'
    elif '\xa3' in fmt_str:
        val = 'GBP'
    elif chr(8364) in fmt_str:
        val = 'EUR'
    elif (fmt_str.find('USD')>=0) or (fmt_str.find('[$$')>=0) :
        val = 'USD'
    else:
        val = ''
    return val



def currencyType(sheet, rowx, colx):
    '''
    Функция анализирует "формат ячейки" таблицы excel, является ли он "денежным"
    и какая валюта указана в этом формате.
    Распознаются не все валюты и способы их описания.
    '''
    c = sheet.cell(rowx, colx)
    xf = sheet.book.xf_list[c.xf_index]
    fmt_obj = sheet.book.format_map[xf.format_key]
    fmt_str = fmt_obj.format_str
    if '\u20bd' in fmt_str:
        val = 'RUB'
    elif '\xa3' in fmt_str:
        val = 'GBP'
    elif chr(8364) in fmt_str:
        val = 'EUR'
    elif (fmt_str.find('USD')>=0) or (fmt_str.find('[$$')>=0) :
        val = 'USD'
    else:
        val = 'RUB'
    return val
    
'''
    return 'usd'

[$$-409]#,##0.0
[$$-409]#,##0.0
[$$-409]#,##0.0
[$$-409]#,##0.0
[$$-409]#,##0.0
#,##0.0"р."
#,##0.0"р."
#,##0.0"р."
#,##0.0"р."
#,##0.0"р."
#
'''



def dump_cell(sheet, rowx, colx):
    c = sheet.cell(rowx, colx)
    xf = sheet.book.xf_list[c.xf_index]
    fmt_obj = sheet.book.format_map[xf.format_key]
    ccc = ord(fmt_obj.format_str[4])
    print( rowx, colx, repr(c.value), c.ctype, fmt_obj.type, ccc, chr(ccc) )
    #print( repr(fmt_obj.format_str))
    


def quoted(sss):
    if ((',' in sss) or ('"' in sss) or ('\n' in sss))  and not(sss[0]=='"' and sss[-1]=='"') :
        sss = '"'+sss.replace('"','""')+'"'
    return sss
